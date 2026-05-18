"""TASK-014: построчная разметка зоны 2 кодификатора ЦУП на AMOS APN / поля /
таблицы / ATA chapter, либо на IATA-732 process / stakeholder, либо на
out_of_scope с обоснованием.

Источник: лист `ЦУП_МЕР_IATA` файла
`input2/ЦУП/Отчетность/Кодификатор_ЦУП_МЕР_IATA_730_732.xlsx`, строки 98..272.

Категории:
  * `section_header` — 10 жирных разделителей (ДКЭ/ЛС/СОП/ПРЧ/ПДГ/НМЧ/М/У/СВС/ППС).
  * `ata_amos`        — кандидат на AMOS-слой: ATA chapter + APN + поле + таблица.
  * `iata732_process` — кандидат на process L1..L2 (G+P) в IATA-732.
  * `iata732_stakeholder` — кандидат на stakeholder/Z в IATA-732.
  * `out_of_scope`    — ни AMOS, ни IATA-732 (admin/IT/security/behaviour/мета).
  * `empty`           — служебная пустая строка 98.

Confidence:
  * `high`   — Hardcoded явная разметка по тексту (или АТА-keyword ≥ 90% уверенно).
  * `medium` — Эвристика по сильному keyword, но возможна вариация.
  * `low`    — Слабый keyword / mixed.
  * `gap`    — Rule 0: не идентифицировано в источнике AMOS Guide.

Rule 0: каждая запись содержит точный `cup_text` (цитата из Excel). Догадки
помечаются `confidence=gap` либо `rule0_note`.

Выход:
  output/amos_layer/cup_zone2_to_amos.json
  output/amos_layer/cup_zone2_to_amos.csv
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path("/home/budnik_an/Obligations")
XLSX = ROOT / "input2" / "ЦУП" / "Отчетность" / "Кодификатор_ЦУП_МЕР_IATA_730_732.xlsx"
SHEET = "ЦУП_МЕР_IATA"
OUT_DIR = ROOT / "output" / "amos_layer"
OUT_JSON = OUT_DIR / "cup_zone2_to_amos.json"
OUT_CSV = OUT_DIR / "cup_zone2_to_amos.csv"
APN_CATALOG = OUT_DIR / "amos_apn_catalog.json"

ROW_FIRST, ROW_LAST = 98, 272


# ---------- справочники ---------------------------------------------------- #

# Стандартные ATA chapters (ASD STE-100 / ATA iSpec 2200), сокращённый список —
# только те, что встречаются в кодификаторе ЦУП НМЧ-секции.
ATA = {
    "12": "Servicing (Routine Maintenance)",
    "21": "Air Conditioning",
    "22": "Auto Flight",
    "23": "Communications",
    "24": "Electrical Power",
    "25": "Equipment / Furnishings",
    "26": "Fire Protection",
    "27": "Flight Controls",
    "28": "Fuel",
    "29": "Hydraulic Power",
    "30": "Ice and Rain Protection",
    "31": "Indicating / Recording Systems",
    "32": "Landing Gear",
    "33": "Lights",
    "34": "Navigation",
    "35": "Oxygen",
    "36": "Pneumatic",
    "38": "Water / Waste",
    "49": "Airborne Auxiliary Power (APU)",
    "50": "Cargo and Accessory Compartments",
    "52": "Doors",
    "53": "Fuselage",
    "54": "Nacelles / Pylons",
    "55": "Stabilizers",
    "56": "Windows",
    "57": "Wings",
    "61": "Propellers / Propulsors",
    "71": "Power Plant — General",
    "73": "Engine Fuel and Control",
    "78": "Engine Exhaust (incl. Thrust Reverser)",
    "80": "Starting",
}

ATA_RU = {
    "12": "Регламентное обслуживание",
    "21": "СКВ / Кондиционирование",
    "22": "Автоматическое управление",
    "23": "Связь и аудио",
    "24": "Электропитание",
    "25": "Оборудование и интерьер",
    "26": "Противопожарная защита",
    "27": "Управление полётом",
    "28": "Топливная система",
    "29": "Гидросистема",
    "30": "Противообледенительная защита",
    "31": "Индикация и регистрация",
    "32": "Шасси",
    "33": "Светотехника",
    "34": "Навигация",
    "35": "Кислородное оборудование",
    "36": "Пневмосистема",
    "38": "Водяная система / сан. узлы",
    "49": "ВСУ",
    "50": "Багажные/грузовые отсеки",
    "52": "Двери",
    "53": "Фюзеляж",
    "54": "Пилоны / гондолы",
    "55": "Стабилизаторы",
    "56": "Остекление",
    "57": "Крыло",
    "61": "Воздушные винты",
    "71": "Силовая установка (общее)",
    "73": "Подача топлива в двигатель",
    "78": "Реверс тяги",
    "80": "Запуск",
}

# Aircraft-centric AMOS lookup chain (Rule: от неисправного ВС обратно к
# причине отказа).  Используется как amos_chain в записях ata_amos.
AMOS_CHAIN = [
    {"step": 1, "node": "aircraft",              "apn": "0308", "table_hint": "ac.aircraft",         "purpose": "вход по ac_registr"},
    {"step": 2, "node": "workorder",             "apn": "1418", "table_hint": "wo.wo_header, jc.jc_header", "purpose": "snag/defect/scheduled"},
    {"step": 3, "node": "ata_chapter",           "apn": "—",    "table_hint": "spec2k.*",            "purpose": "классификация по ATA"},
    {"step": 4, "node": "defect_cause",          "apn": "0354", "table_hint": "rm.failure_*, moc.moc_case", "purpose": "подтверждённая причина"},
    {"step": 5, "node": "deferral_or_close",     "apn": "0273", "table_hint": "mel.mel_*",           "purpose": "MEL (опционально)"},
    {"step": 6, "node": "reliability_aggregation","apn": "0399","table_hint": "rm.*",                "purpose": "агрегация надёжности (опционально)"},
]

# Принятая в проекте дефолтная маршрутизация AMOS-кейсов (см. amos_apn_catalog).
# Дефолт WO применяется как шаг 2 цепочки (после aircraft).
AMOS_DEFAULT_WO = {
    "apn": "1418",
    "apn_name": "Workorder",
    "help_url": "/info/amos-help/APN1418.htm",
    "field": "Defect description / Component / ATA chapter / Task code",
    "table_hint": "wo.wo_header, wo.wo_event_chain, jc.jc_header",
}
AMOS_MEL = {
    "apn": "0273",
    "apn_name": "MEL Manual Administration",
    "help_url": "/info/amos-help/APN0273.htm",
    "field": "MEL code / Cat / DeferralRef / ATA / Open since",
    "table_hint": "mel.mel_*, wo.wo_header (deferral link)",
}
AMOS_TECH_ASSIST = {
    "apn": "0869",
    "apn_name": "Technical Assistance",
    "help_url": "/info/amos-help/APN0869.htm",
    "field": "TQ form U240 / ATA / Aircraft / Status",
    "table_hint": "moc.moc_case, moc.moc_case_log",
}
AMOS_RELIABILITY_SYS = {
    "apn": "0399",
    "apn_name": "Systems Reliability",
    "help_url": "/info/amos-help/APN0399.htm",
    "field": "ATA chapter / Aircraft / Removal events / MTBF",
    "table_hint": "rm.*",
}

# IATA-732 targets (структура задаётся build_structure_json.py):
#   G1..G7 — group, A..N — process, A..N — reason, etc.
# Здесь — короткие ярлыки для разметки. Финальное приклеивание идёт через
# matching.json в следующей итерации.
T732 = {
    # processes (L1+L2)
    "P:G7/Y": "Aircraft defects (нац. отказы — техника)",
    "P:G7/Z": "Aircraft / ATFM technical interruption",
    "P:G1/A": "Aircraft / passenger and baggage flow on ground",
    "P:G2/D": "Aircraft / cargo & mail (loading)",
    "P:G3/E": "Aircraft / fuelling",
    "P:G3/H": "Aircraft / ground handling on stand",
    "P:G4/L": "Aircraft / catering",
    "P:G4/N": "Aircraft / crew (flight operations)",
    "P:G5/P": "Aircraft / passengers (boarding/disembark)",
    "P:G6/W": "Aircraft / weather",
    "P:G7/Z": "Aircraft / ATFM",
    # stakeholders
    "S:A": "Stakeholder A — Airline / handler",
    "S:G": "Stakeholder G — Government / regulator",
    "S:N": "Stakeholder N — Airline crew",
    "S:S": "Stakeholder S — Security",
}


# --------- 10 жирных разделителей зоны 2 (sanity-проверка) ----------------- #
SECTION_HEADERS = {99, 100, 107, 120, 130, 158, 170, 241, 252, 259}


# ---------- Hardcoded по тексту (точная цитата) ---------------------------- #
# Ключ — нормализованный cup_text (lower, без пробелов хвостов).  Значение —
# набор полей разметки.
def _norm(s: str | None) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip())


# AMOS-разметка по точному совпадению cup_text → ATA chapter.
EXPLICIT_ATA_BY_TEXT: dict[str, str] = {
    # НМЧ-секция (стр.171–231 + 192 ЗПС + 224 НМЧ-без-кода)
    "Аварийное освещение": "33",
    "Аварийный трап": "25",
    "Автомат тяги": "22",
    "Автопилот": "22",
    "АКБ": "24",
    "Аудиосистема экипажа": "23",
    "БАНО": "33",
    "Бортовой компьютер/авиационные бортовые приборы/дисплеи": "31",
    "Бортовой самописец": "31",
    "Бытовое оборудование": "25",
    "Водяная система": "38",
    "Воздушный винт (пропеллер)": "61",
    "ВСУ": "49",
    "Гарнитуры/планшеты экипажа": "23",
    "Генераторы": "24",
    "Гидросистема": "29",
    "Датчики угла атаки": "34",
    "Двери входные/аварийные/сервисные/БГО/отсеков/кабины экипажа/аварийный канат": "52",
    "Дозаправка маслом/гидрожидкостью/азотом": "12",
    "Неисправности колес": "32",
    "Замечание по салону ВС/кабине экипажа": "25",
    "Замечание по освещению": "33",
    "Инерциальная система": "34",
    "Кислородное оборудование/АСО": "35",
    "Кресла/привязная система экипажа": "25",
    "Курсовая система": "34",
    "Кухонное оборудование/кофемейкеры": "25",
    "Лобовое стекло/форточки кабины": "56",
    "Мегафон": "23",
    "Метеолокатор/Метеосистема": "34",
    "Механизация крыла": "27",
    "Навигационная система": "34",
    "Неисправность БГО": "50",
    "Неисправность двигателя": "71",
    "Опоры шасси": "32",
    "Пилоны": "54",
    "Посадочно-рулежные фары": "33",
    "Противообледенительная система": "30",
    "Противопожарная система": "26",
    "Радиосвязь/радиооборудование": "23",
    "Реверс двигателя": "78",
    "Руль направления/Руль высоты/Стабилизатор/Триммер/Демпфер рыскания": "27",
    "Сан. узлы": "38",
    "САРД": "21",
    "Система управления": "27",
    "Система внутренней связи/вызова бортпроводников": "23",
    "Система воздушных сигналов": "34",
    "Система регулировки температуры в кабине экипажа": "21",
    "Система уборки/выпуска шасси": "32",
    "Системы обогрева": "30",
    "Системы сигнализации (аварийной, предупреждающей и уведомляющей)": "31",
    "СКВ": "21",
    "Течь гидрожидкости": "29",
    "Течь конденсата в салоне ВС/запотевание иллюминаторов": "21",
    "Топливная система/Панель заправки ВС": "28",
    "Тормозные системы": "32",
    "Фюзеляж/обтекатели": "53",
    "Штурвальная колонка/педали/РУД": "27",
    "Электросистема": "24",
    "Электростатический разрядник": "24",
    # вне-НМЧ AMOS-кандидаты
    "Навигационная база данных": "34",
}

# Спец-роутинг AMOS APN для конкретных текстов (где не WO-default).
EXPLICIT_AMOS_APN: dict[str, dict] = {
    "Выполнение ТО": {  # встречается в стр.184 и 239
        "amos_apn": "1844",
        "amos_apn_name": "Maintenance Forecast",
        "amos_help_url": "/info/amos-help/APN1844.htm",
        "amos_field": "Task / Interval / ATA / Aircraft / Due / Finding",
        "amos_table_hint": "mevt.*, msc.*",
    },
    "Ожидание термосов (открыт п. MEL)": {
        "amos_apn": "0273",
        "amos_apn_name": "MEL Manual Administration",
        "amos_help_url": "/info/amos-help/APN0273.htm",
        "amos_field": "MEL item / Cat / Deferral / linked WO",
        "amos_table_hint": "mel.*, wo.wo_header",
    },
    "Консультация КВС с ИТС": {
        "amos_apn": "0869",
        "amos_apn_name": "Technical Assistance",
        "amos_help_url": "/info/amos-help/APN0869.htm",
        "amos_field": "TQ form U240 / ATA / Aircraft / Status",
        "amos_table_hint": "moc.moc_case, moc.moc_case_log",
    },
    "Заправка ВС при помощи ИТС": {
        "amos_apn": "1418",
        "amos_apn_name": "Workorder",
        "amos_help_url": "/info/amos-help/APN1418.htm",
        "amos_field": "Task code / Servicing / ATA-12 / Aircraft",
        "amos_table_hint": "wo.wo_header (servicing class)",
    },
    "Ожидание доставки компонентов/исправности ВС": {
        "amos_apn": "1208",
        "amos_apn_name": "Shipment Tracking",
        "amos_help_url": "/info/amos-help/APN1208.htm",
        "amos_field": "Shipment / Component / Required date / ATA",
        "amos_table_hint": "od.*, sh.*",
    },
    "Расшифровка полетной информации/оценка состояния ВС при сообщении от экипажа": {
        "amos_apn": "0354",
        "amos_apn_name": "Failure Confirmation",
        "amos_help_url": "/info/amos-help/APN0354.htm",
        "amos_field": "Pilot report / FDR readout / ATA / Aircraft",
        "amos_table_hint": "rm.*, moc.moc_case",
    },
}


# ---------- IATA-732 mapping по тексту ------------------------------------- #
# (target — короткий ярлык; полную геометрию L1+L2/L3 ставим в следующей итерации
# через matching.json).

EXPLICIT_IATA732: dict[str, dict] = {
    # ДКЭ — поведение/готовность бортпроводников
    "Длительный доклад бортпроводников/поздний отгон трапа от бортпроводников": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Замена бортпроводника": {"cat": "iata732_process", "target": "P:G4/N"},
    "Позднее предоставление бортпроводниками информации о замечании по салону ВС": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Ожидание бортпроводников": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Опоздание/неявка бортпроводника": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Поздняя готовность от бортпроводников": {"cat": "iata732_stakeholder", "target": "S:N"},
    # ЛС — лётная служба
    "Летная служба (планирование)": {"cat": "iata732_process", "target": "P:G4/N"},
    "Замена экипажа": {"cat": "iata732_process", "target": "P:G4/N"},
    "Ожидание летного экипажа": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Опоздание/неявка экипажа": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Отдых экипажа": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Отказ экипажа от перевозки опасного груза": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Подготовка экипажа к вылету": {"cat": "iata732_process", "target": "P:G4/N"},
    "Позднее предоставление данных по топливу/неверные данные по топливу": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Поздняя готовность от экипажа": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Урегулирование вопроса по топливу/дозаправка ВС по решению КВС": {"cat": "iata732_process", "target": "P:G3/E"},
    # СОП — пассажиры/багаж
    "Допосадка опоздавших пассажиров": {"cat": "iata732_process", "target": "P:G5/P"},
    "Высадка/посадка маломобильных пассажиров": {"cat": "iata732_process", "target": "P:G5/P"},
    "Несвоевременная доставка пассажиров (позднее предоставление автобусов)": {"cat": "iata732_process", "target": "P:G5/P"},
    "Несвоевременная высадка/посадка пассажиров/Сбой при посадке пассажиров": {"cat": "iata732_process", "target": "P:G5/P"},
    "Несвоевременная регистрация пассажиров/Сбой при регистрации": {"cat": "iata732_process", "target": "P:G1/A"},
    "Ожидание СЗВ/др. док-ов на перевозку (поздняя доставка/перевыпуск/исправление)": {"cat": "iata732_stakeholder", "target": "S:G"},
    "Оплата сверхнормативного багажа/ручной клади": {"cat": "iata732_process", "target": "P:G1/A"},
    "Пересадка пассажиров на борту ВС/снятие пассажира (сдвоенные посадочные талоны)": {"cat": "iata732_process", "target": "P:G5/P"},
    "Разгрузка/загрузка багажа/почты/груза": {"cat": "iata732_process", "target": "P:G2/D"},
    # ПДГ — наземное обслуживание
    "Неисправность спец. техники": {"cat": "iata732_stakeholder", "target": "S:A", "note": "GSE — ground support equipment, не AMOS-flight"},
    "Несвоевременная заправка водяной системы/сан.узлов": {"cat": "iata732_process", "target": "P:G3/H"},
    "Несвоевременная заправка ВС топливом": {"cat": "iata732_process", "target": "P:G3/E"},
    "Несвоевременная подготовка ВС": {"cat": "iata732_process", "target": "P:G3/H"},
    "Несвоевременный заказ бортпитания/ошибка в заказе": {"cat": "iata732_process", "target": "P:G4/L"},
    "Работа агента СПО": {"cat": "iata732_stakeholder", "target": "S:A"},
    "Перебуксировка ВС/замена МС": {"cat": "iata732_process", "target": "P:G3/H"},
    "Поздняя доставка экипажа к месту стоянки ВС": {"cat": "iata732_stakeholder", "target": "S:N"},
    "Поздняя передача готовности ВС": {"cat": "iata732_stakeholder", "target": "S:A"},
    "Поздняя подача/отгон спец.техники": {"cat": "iata732_stakeholder", "target": "S:A"},
    "Недостаточный обогрев ВС": {"cat": "iata732_process", "target": "P:G3/H"},
    # М/У — метеоусловия
    "М/У а/п вылета": {"cat": "iata732_process", "target": "P:G6/W"},
    "М/У а/п прилета": {"cat": "iata732_process", "target": "P:G6/W"},
    "Подготовка ВПП а/п вылета": {"cat": "iata732_process", "target": "P:G6/W"},
    "Подготовка ВПП а/п прилета": {"cat": "iata732_process", "target": "P:G6/W"},
    "Изменение запасных аэродромов": {"cat": "iata732_process", "target": "P:G6/W"},
    "ПОО": {"cat": "iata732_process", "target": "P:G6/W", "note": "противообледенительная обработка"},
    "Отогрев ВС": {"cat": "iata732_process", "target": "P:G6/W"},
    "М/У (прочее)": {"cat": "iata732_process", "target": "P:G6/W"},
    "Ветровой режим": {"cat": "iata732_process", "target": "P:G6/W"},
    "Обдув ВС": {"cat": "iata732_process", "target": "P:G6/W"},
    # СВС — снабжение/уборка
    "уборка ВС": {"cat": "iata732_process", "target": "P:G4/L"},
    "бортовое питание": {"cat": "iata732_process", "target": "P:G4/L"},
    "бытовое имущество": {"cat": "iata732_process", "target": "P:G4/L"},
    "позднее прибытие СС": {"cat": "iata732_stakeholder", "target": "S:A"},
    "перезагрузка БКО и инвентаря": {"cat": "iata732_process", "target": "P:G4/L"},
    "СВС (прочее)": {"cat": "iata732_process", "target": "P:G4/L"},
    # ПРЧ — частично IATA-732, частично out_of_scope
    "Длительная высадка пассажиров по прилету": {"cat": "iata732_process", "target": "P:G5/P"},
    "Длительная доставка пассажиров к МС (занятость маршрута движения автобуса)": {"cat": "iata732_process", "target": "P:G5/P"},
    "Дозаказ бортпитания (повышение в классе обслуживания)": {"cat": "iata732_process", "target": "P:G4/L"},
    "Дополнительная уборка (загрязненный салон/БГО)/уборка ВС экипажем": {"cat": "iata732_process", "target": "P:G4/L"},
    "Допосадка пассажиров по JMP": {"cat": "iata732_process", "target": "P:G5/P"},
    "Ожидание разрешения авиационных властей/утверждение ФПЛ": {"cat": "iata732_stakeholder", "target": "S:G"},
    "Обслуживание ВС при сокращенном ТГО": {"cat": "iata732_process", "target": "P:G3/H"},
    "Размен ВС": {"cat": "iata732_process", "target": "P:G7/Y", "note": "swap борта — обычно следствие AMOS-инцидента"},
    "Регламент а/п": {"cat": "iata732_process", "target": "P:G1/A"},
    "Максимальная взлетная масса": {"cat": "iata732_process", "target": "P:G2/D"},
    # СБОЙ — операционные сбои
    "Запуск от УВЗ/наземного источника питания": {"cat": "iata732_process", "target": "P:G3/H"},
    "Корректировка расписания из-за НМЧ/ЗПС (в базовом а/п)": {"cat": "iata732_process", "target": "P:G7/Y"},
    "Корректировка расписания из-за НМЧ/ЗПС (не в базовом а/п)": {"cat": "iata732_process", "target": "P:G7/Y"},
    # ATFM
    "Закрытие аэропорта (аварийная посадка стороннего ВС/выкатывание за пределы ВПП)": {"cat": "iata732_process", "target": "P:G7/Z"},
    "Изменение слотов (загруженность аэропорта прилета/вылета)": {"cat": "iata732_process", "target": "P:G7/Z"},
}


EXPLICIT_OUT_OF_SCOPE: dict[str, str] = {
    "ЦУП распределения": "мета-метка кодификатора, не строка-кейс",
    "Внесение данных о визах": "административная операция, не процесс задержки",
    "Депортированные пассажиры": "поведенческий/гос. инцидент, частично S:G",
    "Информация о заминировании ВС/аэропорта": "security event, частично S:S, обычно за пределами task delay code",
    "Мед. контроль (противоэпидемические меры)": "санитарные меры — не AMOS, S:G",
    "Мед. мероприятия на борту ВС": "медицинский инцидент в полёте — не AMOS, S:N/S:M",
    "Новая СЗВ/изменения в СЗВ (неявка на посадку)": "admin/документы",
    "Оказание мед. помощи": "медицинский инцидент",
    "Отказ пассажира от полета": "поведенческий, не AMOS",
    "Пограничный контроль": "регуляторный, S:G",
    "Рассадка пассажиров на борту ВС": "operational/seat assignment",
    "Сбой в системе регистрации": "IT/DCS системы, не AMOS",
    "Сбой в системе Интернет": "IT систем а/п, не AMOS",
    "Сбой в системе обработки информации UTG": "IT систем UTG, не AMOS",
    "Нарушение пассажиром правил перевозки/неадекватное состояние/конфликт": "поведенческое",
    "САБ": "служба авиационной безопасности — S:S",
    "СБОЙ (прочее)": "коллективный код-зонтик, прямого AMOS-маппинга нет",
    # ППС-секция (стр.260-272): это деривативные коды («после-причинные сегменты»
    # — что произошло после первичного RZM-кода), не самостоятельные классы.
    "ППС М/У": "post-causal segment ППС, не самостоятельный код",
    "ППС ПОО": "post-causal segment ППС, не самостоятельный код",
    "СБОЙ НМС/ЗПС": "post-causal segment ППС, не самостоятельный код",
    "ППС НО и СВС": "post-causal segment ППС, не самостоятельный код",
    "ППС СБ": "post-causal segment ППС, не самостоятельный код",
    "ППС А/П": "post-causal segment ППС, не самостоятельный код",
    "ППС РЖМ": "post-causal segment ППС, не самостоятельный код",
    "ППС ЛС": "post-causal segment ППС, не самостоятельный код",
    "ППС ДКЭ": "post-causal segment ППС, не самостоятельный код",
    "ППС ПВС": "post-causal segment ППС, не самостоятельный код",
    "ППС ЗАК": "post-causal segment ППС, не самостоятельный код",
    "ППС ОВС": "post-causal segment ППС, не самостоятельный код",
    "ППС ПРОЧЕЕ": "post-causal segment ППС, не самостоятельный код",
}


# ---------- main ----------------------------------------------------------- #


def _chain_for(cup_text: str, ata: str | None) -> list[dict]:
    """Возвращает aircraft-centric цепочку lookup'а для конкретной строки.

    Логика по `EXPLICIT_AMOS_APN` (специальные кейсы):
      * Выполнение ТО (APN 1844)    → steps [1,2,3]      (aircraft → WO → ATA)
      * Ожидание термосов / MEL     → steps [1,2,3,5]    (+MEL)
      * Консультация КВС с ИТС      → steps [1,2,4]      (через Tech Case)
      * Заправка через ИТС          → steps [1,2,3]
      * Ожидание доставки компонент → steps [1,2]        (ВС → WO, shipment отдельно)
      * Расшифровка полётной инф.   → steps [1,2,3,4]    (Failure Confirmation)
    Дефолт (НМЧ snag с ATA): [1,2,3,4].
    """
    special_steps = {
        "Выполнение ТО": [1, 2, 3],
        "Ожидание термосов (открыт п. MEL)": [1, 2, 3, 5],
        "Консультация КВС с ИТС": [1, 2, 4],
        "Заправка ВС при помощи ИТС": [1, 2, 3],
        "Ожидание доставки компонентов/исправности ВС": [1, 2],
        "Расшифровка полетной информации/оценка состояния ВС при сообщении от экипажа": [1, 2, 3, 4],
    }
    steps_idx = special_steps.get(cup_text, [1, 2, 3, 4] if ata else [1, 2])
    return [AMOS_CHAIN[i - 1] for i in steps_idx]


def is_section_header(row: int, cup_text: str, cup_group: str, bold: bool) -> bool:
    if row in SECTION_HEADERS:
        return True
    return bool(bold) and cup_text == cup_group


def classify_row(row: int, cup_text: str, cup_group: str, bold: bool) -> dict:
    record: dict[str, Any] = {
        "row": row,
        "cup_text": cup_text,
        "cup_group": cup_group,
        "is_bold": bold,
        "category": None,
        "ata_chapter": None,
        "ata_chapter_name_en": None,
        "ata_chapter_name_ru": None,
        "amos_apn": None,
        "amos_apn_name": None,
        "amos_help_url": None,
        "amos_field": None,
        "amos_table_hint": None,
        "amos_lookup_chain": None,
        "iata732_target": None,
        "confidence": "gap",
        "rule0_note": "",
    }

    if not cup_text:
        record["category"] = "empty"
        record["confidence"] = "high"
        return record

    if is_section_header(row, cup_text, cup_group, bold):
        record["category"] = "section_header"
        record["confidence"] = "high"
        record["rule0_note"] = "жирный разделитель секции"
        return record

    # 1) AMOS-разметка по точному тексту → ATA chapter
    ata = EXPLICIT_ATA_BY_TEXT.get(cup_text)
    if ata:
        record["category"] = "ata_amos"
        record["ata_chapter"] = ata
        record["ata_chapter_name_en"] = ATA.get(ata)
        record["ata_chapter_name_ru"] = ATA_RU.get(ata)
        # APN — если есть спец-роутинг, иначе дефолт WO 1418
        amos = EXPLICIT_AMOS_APN.get(cup_text)
        if amos:
            record.update(amos)
        else:
            record["amos_apn"] = AMOS_DEFAULT_WO["apn"]
            record["amos_apn_name"] = AMOS_DEFAULT_WO["apn_name"]
            record["amos_help_url"] = AMOS_DEFAULT_WO["help_url"]
            record["amos_field"] = AMOS_DEFAULT_WO["field"]
            record["amos_table_hint"] = AMOS_DEFAULT_WO["table_hint"]
        record["amos_lookup_chain"] = _chain_for(cup_text, ata=ata)
        record["confidence"] = "high"
        return record

    # 2) AMOS-разметка по спец-APN без ATA chapter
    if cup_text in EXPLICIT_AMOS_APN:
        amos = EXPLICIT_AMOS_APN[cup_text]
        record.update(amos)
        record["category"] = "ata_amos"
        record["amos_lookup_chain"] = _chain_for(cup_text, ata=None)
        record["confidence"] = "high"
        record["rule0_note"] = "ATA chapter не однозначен — приписан AMOS-модуль без chapter"
        return record

    # 3) IATA-732 explicit
    if cup_text in EXPLICIT_IATA732:
        spec = EXPLICIT_IATA732[cup_text]
        record["category"] = spec["cat"]
        record["iata732_target"] = spec["target"]
        record["confidence"] = "high"
        if "note" in spec:
            record["rule0_note"] = spec["note"]
        return record

    # 4) Out of scope explicit
    if cup_text in EXPLICIT_OUT_OF_SCOPE:
        record["category"] = "out_of_scope"
        record["confidence"] = "high"
        record["rule0_note"] = EXPLICIT_OUT_OF_SCOPE[cup_text]
        return record

    # 5) GAP — не нашли в источнике / нет точного hardcoded
    record["category"] = "out_of_scope"
    record["confidence"] = "gap"
    record["rule0_note"] = "Rule 0: не идентифицировано в Hardcoded словаре — требуется ручная разметка"
    return record


def main() -> None:
    if not APN_CATALOG.exists():
        raise SystemExit(
            "Сначала запусти scripts/build_amos_apn_catalog.py — "
            f"нет {APN_CATALOG}"
        )

    print(f"reading {XLSX}")
    wb = load_workbook(XLSX, data_only=False)
    ws = wb[SHEET]

    records: list[dict] = []
    for row in range(ROW_FIRST, ROW_LAST + 1):
        a = ws.cell(row=row, column=1).value
        b = ws.cell(row=row, column=2).value
        fa = ws.cell(row=row, column=1).font
        fb = ws.cell(row=row, column=2).font
        bold_a = bool(fa.bold) if a else False
        bold_b = bool(fb.bold) if b else False
        bold = bold_a or bold_b
        cup_text = _norm(a)
        cup_group = _norm(b)

        rec = classify_row(row, cup_text, cup_group, bold)
        records.append(rec)

    # стата
    from collections import Counter
    cat_cnt = Counter(r["category"] for r in records)
    conf_cnt = Counter(r["confidence"] for r in records)
    ata_cnt = Counter(r["ata_chapter"] for r in records if r["ata_chapter"])
    apn_cnt = Counter(r["amos_apn"] for r in records if r["amos_apn"])
    target_cnt = Counter(r["iata732_target"] for r in records if r["iata732_target"])

    print("Категории:", dict(cat_cnt))
    print("Confidence:", dict(conf_cnt))
    print("ATA chapters:", dict(ata_cnt.most_common()))
    print("AMOS APN:", dict(apn_cnt.most_common()))
    print("IATA-732 targets:", dict(target_cnt.most_common()))

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "version": "v1",
        "task": "TASK-014",
        "source": str(XLSX),
        "sheet": SHEET,
        "row_range": [ROW_FIRST, ROW_LAST],
        "amos_aircraft_centric_chain": AMOS_CHAIN,
        "stats": {
            "total_rows": len(records),
            "by_category": dict(cat_cnt),
            "by_confidence": dict(conf_cnt),
            "ata_chapters": dict(ata_cnt.most_common()),
            "amos_apn_usage": dict(apn_cnt.most_common()),
            "iata732_targets": dict(target_cnt.most_common()),
        },
        "records": records,
    }
    OUT_JSON.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"written {OUT_JSON} ({OUT_JSON.stat().st_size} bytes)")

    # CSV для человеческого ревью
    cols = [
        "row", "cup_text", "cup_group", "is_bold",
        "category", "confidence",
        "ata_chapter", "ata_chapter_name_ru",
        "amos_apn", "amos_apn_name", "amos_help_url",
        "amos_field", "amos_table_hint",
        "iata732_target", "rule0_note",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in records:
            w.writerow(r)
    print(f"written {OUT_CSV} ({OUT_CSV.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
