"""
XLSX Extractor - экстрактор для Excel документов

Использует openpyxl для извлечения:
- Таблиц Excel (листы, ячейки, значения)
- Формул и их результатов
- Форматирования ячеек
- Анализ структуры данных

Применение SOLID:
- Single Responsibility: Только обработка XLSX
- Open/Closed: Расширение BaseExtractor
- Dependency Inversion: Зависимость от абстракции

Автор: PDFtoBPMN Project
Дата: 10.11.2025
"""

from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base_extractor import BaseExtractor
from ..models.data_models import TextBlock, TableBlock, BBox


class XLSXExtractor(BaseExtractor):
    """
    Экстрактор для XLSX документов
    
    Ответственность:
    - Извлечение таблиц из листов Excel
    - Парсинг формул и их результатов
    - Определение структуры данных (заголовки, строки)
    - Анализ связей между ячейками
    """
    
    def __init__(self,
                 extract_tables: bool = True,
                 extract_formulas: bool = True,
                 detect_headers: bool = True,
                 data_only: bool = False):
        """
        Инициализация XLSX экстрактора
        
        Args:
            extract_tables: Извлекать таблицы
            extract_formulas: Извлекать формулы (True) или только значения (False)
            detect_headers: Автоматически определять заголовки таблиц
            data_only: Читать только значения (не формулы)
        
        Raises:
            ImportError: Если openpyxl не установлен
        """
        super().__init__(extract_images=False, extract_tables=extract_tables)
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl не установлен. Установите: pip install openpyxl"
            )
        
        self.extract_formulas = extract_formulas
        self.detect_headers = detect_headers
        self.data_only = data_only
    
    def get_supported_extensions(self) -> List[str]:
        """Поддерживаемые расширения: XLSX, XLS"""
        return ['.xlsx', '.XLSX', '.xls', '.XLS']
    
    def extract_document(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Извлечь весь контент из XLSX документа
        
        Args:
            file_path: Путь к XLSX файлу
        
        Returns:
            Список "страниц" (каждый лист Excel = отдельная страница)
        """
        # Валидация файла
        self.validate_file(file_path)
        self.reset_stats()
        
        # Открываем книгу Excel
        workbook = load_workbook(
            filename=file_path,
            data_only=self.data_only,
            read_only=True  # Оптимизация для больших файлов
        )
        
        pages_data = []
        
        # Обрабатываем каждый лист как отдельную "страницу"
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            
            page_data = {
                "page_num": sheet_idx,
                "sheet_name": sheet_name,
                "text_blocks": [],
                "image_blocks": [],
                "table_blocks": [],
                "drawing_blocks": []
            }
            
            # Извлекаем таблицу из листа
            if self.extract_tables:
                table_block = self._extract_sheet_as_table(sheet, sheet_idx)
                if table_block:
                    page_data["table_blocks"].append(table_block)
                    self._stats["table_blocks"] += 1
            
            # Извлекаем метаданные листа как текстовый блок
            meta_text = self._extract_sheet_metadata(sheet, sheet_name)
            if meta_text:
                page_data["text_blocks"].append(meta_text)
                self._stats["text_blocks"] += 1
            
            pages_data.append(page_data)
            self._stats["pages_processed"] += 1
        
        workbook.close()
        
        return pages_data
    
    def _extract_sheet_as_table(self, sheet: Worksheet, sheet_idx: int) -> Optional[TableBlock]:
        """
        Извлечь лист Excel как таблицу
        
        Args:
            sheet: Объект листа openpyxl
            sheet_idx: Индекс листа
        
        Returns:
            TableBlock с данными листа
        """
        # Определяем границы данных
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if max_row == 0 or max_col == 0:
            return None
        
        # Извлекаем данные
        table_data = []
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Формируем значение ячейки
                cell_value = self._format_cell_value(cell)
                row_data.append(cell_value)
            
            table_data.append(row_data)
        
        # Создаем виртуальный BBox
        bbox = BBox(
            x0=0.0,
            y0=0.0,
            x1=float(max_col * 100),  # Виртуальная ширина
            y1=float(max_row * 20)     # Виртуальная высота
        )
        
        return TableBlock(
            bbox=bbox,
            data=table_data,
            page_num=sheet_idx
        )
    
    def _format_cell_value(self, cell: Any) -> str:
        """
        Форматировать значение ячейки с учетом формул
        
        Args:
            cell: Объект ячейки openpyxl
        
        Returns:
            Строковое представление значения
        """
        # Если нужны формулы и они есть
        if self.extract_formulas and hasattr(cell, 'value') and isinstance(cell.value, str):
            if cell.value and cell.value.startswith('='):
                # Возвращаем формулу
                return f"[FORMULA: {cell.value}]"
        
        # Иначе возвращаем значение
        value = cell.value
        
        if value is None:
            return ""
        elif isinstance(value, (int, float)):
            return str(value)
        elif isinstance(value, str):
            return value.strip()
        else:
            return str(value)
    
    def _extract_sheet_metadata(self, sheet: Worksheet, sheet_name: str) -> Optional[TextBlock]:
        """
        Извлечь метаданные листа как текстовый блок
        
        Args:
            sheet: Объект листа
            sheet_name: Название листа
        
        Returns:
            TextBlock с метаданными
        """
        meta_lines = [
            f"# Лист: {sheet_name}",
            f"Размер: {sheet.max_row} строк × {sheet.max_column} столбцов"
        ]
        
        # Добавляем информацию о диапазонах
        if hasattr(sheet, 'tables'):
            if sheet.tables:
                meta_lines.append(f"Именованных таблиц: {len(sheet.tables)}")
        
        meta_text = "\n".join(meta_lines)
        
        # Создаем текстовый блок
        bbox = BBox(x0=0.0, y0=0.0, x1=600.0, y1=50.0)
        
        return TextBlock(
            text=meta_text,
            bbox=bbox,
            font_size=14.0,
            font_name="Calibri",
            is_bold=True,
            is_italic=False,
            page_num=0
        )
    
    def extract_formulas_map(self, file_path: str) -> Dict[str, List[Dict[str, str]]]:
        """
        Дополнительный метод: извлечь карту всех формул в книге
        
        Args:
            file_path: Путь к XLSX файлу
        
        Returns:
            Словарь {sheet_name: [{"cell": "A1", "formula": "=SUM(B1:B10)", "value": "100"}]}
        """
        self.validate_file(file_path)
        
        workbook = load_workbook(filename=file_path, data_only=False)
        formulas_map = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_formulas = []
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        sheet_formulas.append({
                            "cell": cell.coordinate,
                            "formula": cell.value,
                            "value": str(cell.value)  # В режиме data_only=False это формула
                        })
            
            if sheet_formulas:
                formulas_map[sheet_name] = sheet_formulas
        
        workbook.close()
        
        return formulas_map

