"""
XLSX Exporter - генерация Excel таблицы с результатами
"""

from pathlib import Path
from typing import List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from .models import OwnerRecord, ValidationReport


class XLSXExporter:
    """Экспорт записей владельцев в Excel"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def export(self, records: List[OwnerRecord], output_path: Path) -> ValidationReport:
        """
        Экспорт записей в XLSX файл
        
        Args:
            records: Список OwnerRecord
            output_path: Путь к выходному файлу
            
        Returns:
            ValidationReport с статистикой
        """
        # Создаем workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Владельцы"
        
        # Заголовки (2 ключевых поля + доп. информация)
        headers = [
            'Адрес регистрации',
            'Количество в штуках',
            'Код владельца',
            'ФИО',
            'Номер документа',
            'Номер счета',
            'Страница'
        ]
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 60  # Адрес
        ws.column_dimensions['B'].width = 20  # Количество
        ws.column_dimensions['C'].width = 20  # Код
        ws.column_dimensions['D'].width = 35  # ФИО
        ws.column_dimensions['E'].width = 18  # Номер документа
        ws.column_dimensions['F'].width = 18  # Счет
        ws.column_dimensions['G'].width = 10  # Страница
        
        # Высота строки заголовка
        ws.row_dimensions[1].height = 30
        
        # Записываем данные
        report = ValidationReport()
        
        for row_idx, record in enumerate(records, 2):
            data = record.to_dict()
            
            ws.cell(row=row_idx, column=1, value=data['Адрес регистрации'])
            ws.cell(row=row_idx, column=2, value=data['Количество в штуках'])
            ws.cell(row=row_idx, column=3, value=data['Код владельца'])
            ws.cell(row=row_idx, column=4, value=data['ФИО'])
            ws.cell(row=row_idx, column=5, value=data['Номер документа'])
            ws.cell(row=row_idx, column=6, value=data['Номер счета'])
            ws.cell(row=row_idx, column=7, value=data['Страница'])
            
            # Выравнивание
            ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center")
            
            # Подсвечиваем невалидные записи (только 2 ключевых поля)
            if not record.validate():
                for col_idx in range(1, 3):
                    ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )
            
            # Добавляем в отчет
            report.add_record(record)
        
        # Создаем лист со статистикой
        self._add_stats_sheet(wb, report)
        
        # Сохраняем
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return report
    
    def _add_stats_sheet(self, wb: openpyxl.Workbook, report: ValidationReport):
        """Добавляет лист со статистикой"""
        ws = wb.create_sheet("Статистика")
        
        # Заголовок
        ws['A1'] = "Отчет валидации"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Статистика
        row = 3
        stats = [
            ("Всего записей:", report.total_records),
            ("Валидных записей:", report.valid_records),
            ("Невалидных записей:", report.invalid_records),
            ("", ""),
            ("Всего бумаг (шт):", report.total_quantity),
            ("", ""),
            ("Проблемы:", ""),
            ("  Отсутствует адрес:", report.missing_address),
            ("  Отсутствует количество:", report.missing_quantity),
        ]
        
        for label, value in stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            
            if "Проблемы" in label or "Всего бумаг" in label:
                ws.cell(row=row, column=1).font = Font(bold=True)
            
            row += 1
        
        # Качество
        quality = (report.valid_records / report.total_records * 100) if report.total_records > 0 else 0
        row += 1
        ws.cell(row=row, column=1, value="Качество данных:")
        ws.cell(row=row, column=2, value=f"{quality:.1f}%")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True, color="008000" if quality >= 95 else "FF0000")
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

